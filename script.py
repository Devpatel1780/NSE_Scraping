from selenium import webdriver
import pandas as pd
import csv
import time
import datetime
import openpyxl
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

today = datetime.date.today()
birth = datetime.date(2022, 3, 25)
diff = today - birth
k = int(diff.days)
print(k)

s = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=s)

agla = str('https://www1.nseindia.com/content/nsccl/')
linkOI = agla + 'fao_participant_oi_' + str(datetime.datetime.now().strftime('%d%m%Y')) + '.csv'
linkVO = agla + 'fao_participant_vol_' + str(datetime.datetime.now().strftime('%d%m%Y')) + '.csv'
driver.get(linkOI)
driver.get(linkVO)
time.sleep(5)

agla2 = str('C:\\Users\\DeLL\\Downloads\\')
fileOI = agla2 + 'fao_participant_oi_' + str(datetime.datetime.now().strftime('%d%m%Y')) + '.csv'

with open(fileOI) as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    OIdf = pd.DataFrame([csv_reader], index=None)
    data = OIdf.head()

a = 2
sno = 0
filename1 = ["OI_Client", "OI_DII", "OI_FII", 'OI_Pro']
exel_file = openpyxl.load_workbook('Volume data sheet.xlsx')
while a < 6:
    for value in list(OIdf[a]):
        cal = 1
        while cal < 15:
            exel_sheet = exel_file[filename1[sno]]
            OI_data = value[cal]
            cal = cal + 1
            time.sleep(2)
            exel_sheet.cell(row=k, column=cal).value = OI_data
    sno += 1
    a += 1

fileVO = agla2 + 'fao_participant_vol_' + str(datetime.datetime.now().strftime('%d%m%Y')) + '.csv'
with open(fileVO) as csv_file1:
    csv_reader1 = csv.reader(csv_file1, delimiter=',')
    VOdf = pd.DataFrame([csv_reader1], index=None)
    data1 = VOdf.head()

e = 2
vno = 0
filename2 = ["VO_Client", 'VO_DII', "VO_FII", "VO_Pro"]
while e < 6:
    for value in list(VOdf[e]):
        cal1 = 1
        while cal1 < 15:
            exel_sheet1 = exel_file[filename2[vno]]
            VO_data = value[cal1]
            cal1 = cal1 + 1
            time.sleep(2)
            exel_sheet1.cell(row=k, column=cal1).value = int(VO_data)
    vno += 1
    e += 1

exel_file.save('Volume data sheet.xlsx')
